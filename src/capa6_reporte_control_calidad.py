#!/usr/bin/env python3
"""Capa 6 - Reporte de control de calidad (Excel).

Genera `reporte_qa.xlsx` a partir de `validacion_resultados.csv` y, opcionalmente,
`ingesta_manifest.csv` para trazabilidad.
"""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path


@dataclass
class ValidationResult:
    archivo_cap7: str
    pagina_cap7: int
    parrafo_idx: int
    tipo: str
    referencia_original: str
    capitulo_objetivo: str
    id_normalizado_ref: str
    estado: str
    detalle: str
    id_match: str
    titulo_match: str
    pagina_match: int
    similitud_titulo: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Capa 6: reporte de control de calidad")
    parser.add_argument(
        "--validacion-csv",
        type=Path,
        default=Path("out/validacion_resultados.csv"),
        help="Archivo CSV de salida de Capa 5",
    )
    parser.add_argument(
        "--manifest-csv",
        type=Path,
        default=None,
        help="Archivo CSV de ingesta (opcional) para hoja de trazabilidad",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("out/reporte_qa.xlsx"),
        help="Ruta de salida del reporte Excel",
    )
    return parser.parse_args()


def _to_int(v: str | int | None) -> int:
    try:
        return int(v or 0)
    except ValueError:
        return 0


def _to_float(v: str | float | None) -> float:
    try:
        return float(v or 0.0)
    except ValueError:
        return 0.0


def load_validacion(path: Path) -> list[ValidationResult]:
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"CSV de validación no encontrado: {path}")

    out: list[ValidationResult] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            out.append(
                ValidationResult(
                    archivo_cap7=(row.get("archivo_cap7") or "").strip(),
                    pagina_cap7=_to_int(row.get("pagina_cap7")),
                    parrafo_idx=_to_int(row.get("parrafo_idx")),
                    tipo=(row.get("tipo") or "").strip(),
                    referencia_original=(row.get("referencia_original") or "").strip(),
                    capitulo_objetivo=(row.get("capitulo_objetivo") or "").strip(),
                    id_normalizado_ref=(row.get("id_normalizado_ref") or "").strip(),
                    estado=(row.get("estado") or "").strip(),
                    detalle=(row.get("detalle") or "").strip(),
                    id_match=(row.get("id_match") or "").strip(),
                    titulo_match=(row.get("titulo_match") or "").strip(),
                    pagina_match=_to_int(row.get("pagina_match")),
                    similitud_titulo=_to_float(row.get("similitud_titulo")),
                )
            )
    return out


def load_manifest_rows(path: Path | None) -> list[dict[str, str]]:
    if path is None:
        return []
    if not path.exists() or not path.is_file():
        return []

    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def build_resumen(rows: list[ValidationResult]) -> list[tuple[str, str, int]]:
    by_estado = Counter(r.estado for r in rows)
    by_capitulo = Counter(r.capitulo_objetivo for r in rows)

    out: list[tuple[str, str, int]] = []
    for k, v in sorted(by_estado.items()):
        out.append(("estado", k, v))
    for k, v in sorted(by_capitulo.items()):
        out.append(("capitulo_objetivo", k, v))
    return out


def build_renumeraciones(rows: list[ValidationResult]) -> list[ValidationResult]:
    return [r for r in rows if r.estado == "REVISAR_RENUMERACION"]


def create_excel(output_path: Path, rows: list[ValidationResult], manifest_rows: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Dependencia faltante: instala `openpyxl` (pip install openpyxl)") from exc

    wb = Workbook()

    ws_resumen = wb.active
    ws_resumen.title = "resumen"
    ws_resumen.append(["dimension", "clave", "cantidad"])
    for dim, key, val in build_resumen(rows):
        ws_resumen.append([dim, key, val])

    ws_detalle = wb.create_sheet("detalle")
    ws_detalle.append(
        [
            "archivo_cap7",
            "pagina_cap7",
            "parrafo_idx",
            "tipo",
            "referencia_original",
            "capitulo_objetivo",
            "id_normalizado_ref",
            "estado",
            "detalle",
            "id_match",
            "titulo_match",
            "pagina_match",
            "similitud_titulo",
        ]
    )
    for r in rows:
        ws_detalle.append(
            [
                r.archivo_cap7,
                r.pagina_cap7,
                r.parrafo_idx,
                r.tipo,
                r.referencia_original,
                r.capitulo_objetivo,
                r.id_normalizado_ref,
                r.estado,
                r.detalle,
                r.id_match,
                r.titulo_match,
                r.pagina_match,
                r.similitud_titulo,
            ]
        )

    ws_ren = wb.create_sheet("renumeraciones_sugeridas")
    ws_ren.append(
        [
            "tipo",
            "capitulo_objetivo",
            "id_referencia_cap7",
            "id_sugerido",
            "titulo_sugerido",
            "pagina_sugerida",
            "similitud_titulo",
        ]
    )
    for r in build_renumeraciones(rows):
        ws_ren.append(
            [
                r.tipo,
                r.capitulo_objetivo,
                r.id_normalizado_ref,
                r.id_match,
                r.titulo_match,
                r.pagina_match,
                r.similitud_titulo,
            ]
        )

    ws_trz = wb.create_sheet("trazabilidad")
    ws_trz.append(["campo", "valor"])
    ws_trz.append(["generado_en", datetime.now().isoformat(timespec="seconds")])
    ws_trz.append(["total_referencias_validadas", len(rows)])
    if manifest_rows:
        first = manifest_rows[0]
        ws_trz.append(["lote", first.get("lote", "")])
        ws_trz.append(["archivos_ingestados", len(manifest_rows)])
        ws_trz.append([])
        ws_trz.append(["manifest_columns", ", ".join(first.keys())])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()

    try:
        rows = load_validacion(args.validacion_csv.resolve())
    except FileNotFoundError as exc:
        print(f"ERROR: {exc}")
        return 2

    manifest_rows = load_manifest_rows(args.manifest_csv.resolve()) if args.manifest_csv else []

    try:
        create_excel(args.output_xlsx.resolve(), rows, manifest_rows)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return 2

    print(f"Reporte generado: {args.output_xlsx.resolve()}")
    print(f"Total registros en detalle: {len(rows)}")
    print(f"Renumeraciones sugeridas: {len(build_renumeraciones(rows))}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
